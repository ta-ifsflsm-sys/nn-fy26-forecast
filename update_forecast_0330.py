#!/usr/bin/env python3
"""
IFS Japan - Weekly Forecast Update Script
==========================================
使い方:
  python update_forecast.py <sales_opp.xlsx> <previous_forecast.xlsx>
"""

import openpyxl
from datetime import datetime
import re
import sys
import os


def normalize_call_type(value):
    if value is None:
        return None
    return re.sub(r'\s+', ' ', str(value).strip())


def parse_date(d):
    if d is None:
        return None
    if isinstance(d, datetime):
        return d
    if isinstance(d, str):
        for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%Y/%m/%d'):
            try:
                return datetime.strptime(d, fmt)
            except ValueError:
                continue
    return None


def get_quarter(date):
    if date is None:
        return None
    m = date.month
    if m <= 3:  return 1
    if m <= 6:  return 2
    if m <= 9:  return 3
    return 4


def col_letter(n):
    result = ''
    while n:
        n, r = divmod(n - 1, 26)
        result = chr(65 + r) + result
    return result


def detect_header_cols(ws, row=1):
    return {cell.value: cell.column for cell in ws[row] if cell.value is not None}


def update_forecast(source_file, forecast_file):
    today_str = datetime.now().strftime('%Y%m%d_%H%M')
    output_file = f'FY26_Forecast_Call_{today_str}.xlsx'

    print(f"\n{'='*55}")
    print(f"  IFS Japan Weekly Forecast Updater")
    print(f"{'='*55}")
    print(f"  ソース      : {os.path.basename(source_file)}")
    print(f"  前週ファイル: {os.path.basename(forecast_file)}")
    print(f"  出力        : {output_file}")
    print(f"{'='*55}\n")

    # STEP 1: 前週 Last Wk 値を取得
    print("[1/5] 前週データを読み込み中...")
    wb_old    = openpyxl.load_workbook(forecast_file)
    ws_fc_old = wb_old['FY26 Forecast (Call)']
    fc_h      = detect_header_cols(ws_fc_old)

    c_type = fc_h.get('Type', 1)
    c_acv  = fc_h.get('ACV', 4)
    c_lic  = fc_h.get('License Value', 5)
    c_date = fc_h.get('Close Date', 8)

    old_rows = []
    for row in ws_fc_old.iter_rows(min_row=2, values_only=True):
        if not any(v is not None for v in row):
            continue
        old_rows.append({
            'call_type':   normalize_call_type(row[c_type - 1]),
            'acv':         row[c_acv - 1] or 0,
            'license_val': row[c_lic - 1] or 0,
            'sign_date':   parse_date(row[c_date - 1]),
        })

    last_wk = {}
    for q in range(1, 5):
        q_rows = [r for r in old_rows
                  if get_quarter(r['sign_date']) == q and r['call_type'] == 'In Call']
        last_wk[q] = (
            sum(r['acv']         for r in q_rows),
            sum(r['license_val'] for r in q_rows),
        )
        print(f"   Last Wk Q{q}: ACV={last_wk[q][0]:>12,.0f}  License={last_wk[q][1]:>12,.0f}")

    # STEP 2: Sales Opportunities 読み込み
    print("\n[2/5] Sales Opportunities 読み込み中...")
    wb_src  = openpyxl.load_workbook(source_file, data_only=True)
    ws_src  = wb_src.active
    src_h   = {cell.value: i for i, cell in enumerate(ws_src[1])}

    def gc(*names):
        for n in names:
            if n in src_h:
                return src_h[n]
        return None

    idx_desc  = gc('Opp.Description', 'Opp. Description')
    idx_id    = gc('Opportunity ID')
    idx_acv   = gc('Total ACV')
    idx_lic   = gc('Total License')
    idx_book  = gc('New Bookings')
    idx_ai    = gc('AI Bookings', 'AI Booking')
    idx_date  = gc('Sign Date')
    idx_type  = gc('Call Type')

    new_rows = []
    skipped  = 0
    for row in ws_src.iter_rows(min_row=2, values_only=True):
        opp_id = row[idx_id] if idx_id is not None else None
        if not opp_id or str(opp_id).strip() == '':
            skipped += 1
            continue
        ai_val = row[idx_ai] if idx_ai is not None else 0
        if not isinstance(ai_val, (int, float)):
            ai_val = 0

        new_rows.append({
            'call_type':   normalize_call_type(row[idx_type]),
            'opp_desc':    row[idx_desc],
            'opp_id':      str(opp_id),
            'acv':         row[idx_acv]  or 0,
            'license_val': row[idx_lic]  or 0,
            'booking_val': row[idx_book] or 0,
            'ai_booking':  ai_val,
            'sign_date':   parse_date(row[idx_date]),
        })

    new_rows.sort(key=lambda r: r['sign_date'] or datetime(9999, 12, 31))

    quarters = {1: [], 2: [], 3: [], 4: []}
    for r in new_rows:
        q = get_quarter(r['sign_date'])
        if q:
            quarters[q].append(r)

    total_deals = sum(len(v) for v in quarters.values())
    print(f"   読み込み件数: {total_deals}件  (スキップ: {skipped}行)")
    for q in range(1, 5):
        in_c  = sum(1 for r in quarters[q] if r['call_type'] == 'In Call')
        out_c = sum(1 for r in quarters[q] if r['call_type'] == 'Out Call')
        other = len(quarters[q]) - in_c - out_c
        print(f"   Q{q}: {len(quarters[q])}件 (In Call={in_c}, Out Call={out_c}"
              + (f", その他={other}" if other else "") + ")")

    # STEP 3: FY26 Forecast (Call) シートを更新
    print("\n[3/5] FY26 Forecast (Call) シートを更新中...")
    wb    = openpyxl.load_workbook(forecast_file)
    ws_fc = wb['FY26 Forecast (Call)']
    fc_h2 = detect_header_cols(ws_fc)

    wc_type = fc_h2.get('Type', 1)
    wc_desc = fc_h2.get('Opp. Description', 2)
    wc_id   = fc_h2.get('Opportunity No', 3)
    wc_acv  = fc_h2.get('ACV', 4)
    wc_lic  = fc_h2.get('License Value', 5)
    wc_book = fc_h2.get('Booking Value', 6)
    wc_ai   = fc_h2.get('AI Booking', 7)
    wc_date = fc_h2.get('Close Date', 8)

    for row in ws_fc.iter_rows(min_row=2, max_row=ws_fc.max_row):
        for cell in row:
            cell.value = None

    current_row = 2
    quarter_row_ranges = {}

    for q in range(1, 5):
        start = current_row
        for r in quarters[q]:
            ws_fc.cell(current_row, wc_type).value  = r['call_type']
            ws_fc.cell(current_row, wc_desc).value  = r['opp_desc']
            ws_fc.cell(current_row, wc_id).value    = r['opp_id']
            ws_fc.cell(current_row, wc_acv).value   = r['acv']
            ws_fc.cell(current_row, wc_lic).value   = r['license_val']
            ws_fc.cell(current_row, wc_book).value  = r['booking_val']
            ws_fc.cell(current_row, wc_ai).value    = r['ai_booking'] if r['ai_booking'] else None
            d = r['sign_date']
            ws_fc.cell(current_row, wc_date).value  = d.strftime('%Y-%m-%d') if d else None
            current_row += 1
        if not quarters[q]:
            current_row += 1
        quarter_row_ranges[q] = (start, current_row - 1)
        print(f"   Q{q} 行範囲: {quarter_row_ranges[q][0]} - {quarter_row_ranges[q][1]}  ({len(quarters[q])}件)")

    # STEP 4: Dashboard SUMIF 数式を更新
    print("\n[4/5] Dashboard の数式を更新中...")
    ws_dash = wb['Dashboard']
    dash_h  = detect_header_cols(ws_dash, row=1)

    dc_acv  = dash_h.get('ACV', 4)
    dc_lic  = dash_h.get('License', 5)
    dc_book = dash_h.get('Booking', 6)
    dc_ai   = dash_h.get('AI Booking', 7)

    FC = "'FY26 Forecast (Call)'"
    dash_rows = {1: (2, 3), 2: (4, 5), 3: (6, 7), 4: (8, 9)}

    for q in range(1, 5):
        sr, er          = quarter_row_ranges[q]
        in_row, out_row = dash_rows[q]
        ref             = f"$A${sr}:$A${er}"
        for dc, wc in [(dc_acv, wc_acv), (dc_lic, wc_lic),
                       (dc_book, wc_book), (dc_ai, wc_ai)]:
            if dc is None:
                continue
            fc_l       = col_letter(wc)
            data_range = f"{fc_l}{sr}:{fc_l}{er}"
            ws_dash.cell(in_row,  dc).value = (
                f'=SUMIF({FC}!{ref},"In Call",{FC}!{data_range})')
            ws_dash.cell(out_row, dc).value = (
                f'=SUMIF({FC}!{ref},"Out Call",{FC}!{data_range})')

    # STEP 5: Last Wk / WTW を更新
    print("[5/5] Last Wk / WTW を設定中...")

    lw_q_rows = {}
    for r_idx in range(1, ws_dash.max_row + 1):
        for c_idx in range(1, ws_dash.max_column + 1):
            v = ws_dash.cell(r_idx, c_idx).value
            if v in ('Q1', 'Q2', 'Q3', 'Q4') and r_idx > 15:
                q_num = int(str(v)[1])
                lw_q_rows[q_num] = (r_idx, c_idx + 1, c_idx + 2)

    for q in range(1, 5):
        if q in lw_q_rows:
            r, ca, cl = lw_q_rows[q]
            ws_dash.cell(r, ca).value = last_wk[q][0]
            ws_dash.cell(r, cl).value = last_wk[q][1]
            print(f"   Last Wk Q{q}: 行{r} ACV={last_wk[q][0]:,.0f}  Lic={last_wk[q][1]:,.0f}")

    in_call_dash_rows = {q: dash_rows[q][0] for q in range(1, 5)}
    for q in range(1, 5):
        if q not in lw_q_rows:
            continue
        lw_r, lw_ca, lw_cl = lw_q_rows[q]
        in_r = in_call_dash_rows[q]
        wtw_r = None
        for rr in range(1, 16):
            v = ws_dash.cell(rr, lw_ca).value
            if isinstance(v, str) and f'D{in_r}' in v:
                wtw_r = rr
                break
        if wtw_r:
            ws_dash.cell(wtw_r, lw_ca).value = f'=D{in_r}-{col_letter(lw_ca)}{lw_r}'
            ws_dash.cell(wtw_r, lw_cl).value = f'=E{in_r}-{col_letter(lw_cl)}{lw_r}'

    print("   WTW 数式を更新しました")

    wb.save(output_file)

    print(f"\n{'='*55}")
    print(f"  ✅ 完了: {output_file}")
    print(f"{'='*55}")
    print("\n📌 次のステップ:")
    print("  1. 出力ファイルを Excel で開く")
    print("  2. 「編集を有効にする」をクリック → 自動再計算")
    print("  3. 数字がおかしければ Ctrl+Alt+F9 で強制再計算")
    print("  4. Dashboard を確認してレポート完了\n")

    return output_file


if __name__ == '__main__':
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)
    source_file, forecast_file = sys.argv[1], sys.argv[2]
    for f in [source_file, forecast_file]:
        if not os.path.exists(f):
            print(f"❌ エラー: ファイルが見つかりません → {f}")
            sys.exit(1)
    update_forecast(source_file, forecast_file)
